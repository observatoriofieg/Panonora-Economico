#bibliotecas 
from logzero import logger
import subprocess
from upload import *
from github import Github
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import traceback
import logging
import utils
from upload_json_to_github import *


import sys
###
from fgv.panorama_fgv_icc import *
from fgv.panorama_fgv_igpm import *
from ibge.panorama_ibge_inpc import *
from ibge.panorama_ibge_ipca import *
from ibge.panorama_ibge_pim import *
from ibge.panorama_ibge_taxa_desocupacao import *
from me.panorama_me_comex import *
from mte.panorama_mte_saldo_empregados import *
from bacen.panorama_bacen_ibc import *
from cni.panorama_cni_confianca_industrial import *
from cni.panorama_cni_estoque_efetivo import *
from cni.panorama_cni_intencao_investir import *
from cni.panorama_cni_perspectiva_emprego import *
from cni.panorama_cni_utilizacao_capacidade_instalada import *

def recupera_datas(num_sheet):
    caminho_planilha = 'G:/IEL/OBSERVATORIO/ETL/PANORAMA ECONOMICO/calendario de dados panorama economico wendel.xlsx'
    wb = load_workbook(filename=caminho_planilha)
    sheet_name = wb.sheetnames[num_sheet]
    ws = wb[sheet_name]
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df.iloc[1:]
    df.dropna(inplace=True)

    return df

def main():
    data_atual = datetime.now()
    mes_atual = data_atual.month
    num_sheet = mes_atual
    df_calendario = recupera_datas(num_sheet)

    data_atual = datetime.strftime(data_atual, '%Y/%m/%d')
    data_atual = datetime.strptime(data_atual, '%Y/%m/%d')
    print(type(data_atual))

    lista_de_funcoes = [
        {'id': '#1', 'funcao': extracao_ibc, 'json_name': 'panorama_bacen_ibc'},
        {'id': '#2', 'funcao': extracao_confianca_industrial, 'json_name': 'panorama_cni_confianca_industrial'},
        {'id': '#3', 'funcao': extracao_estoque_efetivo, 'json_name': 'panorama_cni_estoque_efetivo'},
        {'id': '#4', 'funcao': extracao_intencao_investir, 'json_name': 'panorama_cni_intencao_investir'},
        {'id': '#5', 'funcao': extracao_perspectiva_emprego, 'json_name': 'panorama_cni_perspectiva_emprego'},
        {'id': '#6', 'funcao': extracao_utilizacao_capacidade_instalada, 'json_name': 'panorama_cni_utilizacao_capacidade_instalada'},
        {'id': '#7', 'funcao': extracao_inpc, 'json_name': 'panorama_ibge_inpc'},
        {'id': '#8', 'funcao': extracao_ipca, 'json_name': 'panorama_ibge_ipca'},
        {'id': '#9', 'funcao': extracao_pim, 'json_name': 'panorama_ibge_pim'},
        {'id': '#10', 'funcao': extracao_taxa_desocupacao, 'json_name': 'panorama_ibge_taxa_desocupacao'},
        {'id': '#11', 'funcao': extracao_comex, 'json_name': '##<##'},
        {'id': '#12', 'funcao': extracao_saldo_empregados, 'json_name':'panorama_mte_saldo_empregados'},
        {'id': '#13', 'funcao': extracao_icc, 'json_name': 'panorama_fgv_icc'},
        {'id': '#14', 'funcao': extracao_igpm, 'json_name': 'panorama_fgv_igpm'},
    ]

    lista_para_atualizacao = []

    for index, row in df_calendario.iterrows():        
        if row[2] <= data_atual  and data_atual <= row[3]:
            for func in zip(lista_de_funcoes):
                func = list(func)[0]
                if func['id'] == row[0]:                
                    lista_para_atualizacao.append({'id': row[0], 'dado': row[1], 'funcao': func['funcao'], 
                                                'json_name': func['json_name']})

    # print(lista_para_atualizacao)
    path_json = utils.config['path_save_json']['path']
    
    qtd_atualizacoes = len(lista_para_atualizacao)
    print(qtd_atualizacoes)
    while qtd_atualizacoes >= 0:
        for index, dado_para_atualizar in enumerate(lista_para_atualizacao):
            
            try:
                json_extraido = dado_para_atualizar['funcao']()
                name_json = dado_para_atualizar['json_name']

                with open(path_json + name_json + ".json", "r", encoding="utf-8") as fp:
                    json_armazenado = json.load(fp)

                if json_extraido != json_armazenado:
                    print('#########', dado_para_atualizar['dado'])
                     
                    del lista_para_atualizacao[index]
                    qtd_atualizacoes =- 1
                    upload_files_to_github(name_json)

                    #ENVIAR EMAIL OU SMS

            except Exception as e: # catch *all* exceptions

                #CRIAR UM LOG

                e = sys.exc_info()[0]
                logging.error(traceback.format_exc())
                print(e)
                pass

            qtd_atualizacoes =- 1

if __name__ == "__main__":
    
    logger.error('*' * 80)
    logger.error('Inicialização do aplicativo')
    logger.error('*' * 80)

    main()

    logger.error('*' * 80)
    logger.error('Finalização do aplicativo')
    logger.error('*' * 80)

    #OK#puxar datas
        #   [{dado: ibc, 'data_inicio': 14/04/2021, 'data_termino': 16,04/2021, indicador_atualizacao: '#'}]
    #OK #verifica quais podem atualziar no dia no dia em questao
    
    #suar TRY 
    #roda o for ate que seja atualizado
        #se atualziado, envia para o github 
        # armazenar qual foi att e qual dado numa lista e qdn consluir o for, enviar email


    #AQUI PODE EXECUTADO UMA FUNCAO PARA PEGAR AS DATAS E HORARIOS DE ATUALIZACAO
    #OU PODE TER UMA FUNCAO EM UTILS QUE FAZ ESSA TAREFA E ENTAO COLOCAR EM CADA UM 
    #DOS .PY E RODAR O FOR, QUANDO DAIR DO FOR, O RESPECTIVO .BAT TERA UM move nul 2 >&0



    # COLOCAR CAMPO JSON PARA INDICAR A DATA DA ULTIMA ATT
        #PARA OS QUE TEM DUAS DATAS OU SE TIVER UMA PERGAR DA DATA INDICADA ATE A DATA ATUAL
        #    SE DATA ATUAL FOR MAIOR QUE DATA DE ATUALIZAÇÃO E DATA ATUALIZAÇÃO ESTIVER ENTRE ESTES DOIS INTERVALOS,
                    # ENTAO JA FOI ATUALIZADO

    #enviar email sms
    ################
    ################ DEFINIAR QUANTAS VEZES ESTA FUNCAO IRA RODAR



    #FECHAR JANELA SE N HOUVER POSSIBILDIADE DAQUELE INDICADOR SER ATUALZIADO NAQUELE DIA


    # #BACEN
    # subprocess.call([r'G:/IEL/OBSERVATORIO/ETL/PANORAMA ECONOMICO/SCRIPTS/BACEN/panorama_bacen_ibc.bat'])
